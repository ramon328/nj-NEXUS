#!/usr/bin/env python3
# mallorca.py — Conector del Excel global de Mallorca para Nexus (solo LECTURA).
#
# POR QUÉ: Mallorca lleva en un Excel de OneDrive los datos financieros del negocio
# (stock valorizado con COSTO/GASTOS/TOTAL, ventas con MARGEN, compras, etc.). GoAutos
# tiene el catálogo/publicación pero NO el costo ni el margen. La llave de cruce es la
# PATENTE. Este conector descarga el Excel (cacheado), lo lee y entrega JSON para que
# el cerebro complemente la info de GoAutos (margen por auto, stock valorizado, ventas).
#
# Uso (devuelve JSON):
#   python mallorca.py hojas                      (mapa del libro: hojas + columnas)
#   python mallorca.py hoja --nombre "VENTAS DE AUTOS" [--limite 30] [--buscar texto]
#   python mallorca.py auto --patente RGVG27      (costo/gastos/total/PV + ventas/compras del auto)
#   python mallorca.py stock                       (stock valorizado: total invertido + lista)
#   python mallorca.py ventas [--mes 2026-06]      (ventas y márgenes; opcional por mes)
#   python mallorca.py refrescar                   (fuerza re-descarga del Excel)
#
# El link del Excel está en ~/nexus/credenciales.json -> mallorca_excel.url
# (link de compartir de OneDrive; si caduca, reemplazarlo). Cache TTL 6 min:
# se re-descarga el Excel si la caché tiene más de 6 minutos, y SIEMPRE se conserva
# solo el archivo más reciente (el anterior se borra) para no acumular almacenamiento.

import sys, os, json, re, argparse, datetime, urllib.request, http.cookiejar

BASE = os.path.dirname(os.path.abspath(__file__))
CACHE = os.path.join(BASE, "cache", "mallorca.xlsx")
CREDS = os.path.join(BASE, "..", "credenciales.json")
TTL_SEG = 6 * 60  # 6 minutos: se re-descarga si la caché tiene más de 6 min
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")

# Hojas de AUTOS (las que tienen PATENTE como llave de cruce con GoAutos).
HOJA_STOCK = "STOCK VALORIZADO"
HOJA_VENTAS = "VENTAS DE AUTOS"
HOJA_COMPRAS = "COMPRAS DE AUTOS"


def salir(obj):
    print(json.dumps(obj, ensure_ascii=False, default=str, indent=2))
    sys.exit(0)


def url_excel():
    try:
        with open(CREDS, "r", encoding="utf-8") as f:
            return (json.load(f).get("mallorca_excel") or {}).get("url") or ""
    except Exception:
        return ""


def descargar(forzar=False):
    """Descarga el Excel a CACHE si no existe o está vencido (o si forzar)."""
    if (not forzar) and os.path.exists(CACHE):
        edad = datetime.datetime.now().timestamp() - os.path.getmtime(CACHE)
        if edad < TTL_SEG:
            return CACHE
    url = url_excel()
    if not url:
        raise RuntimeError("Falta mallorca_excel.url en credenciales.json")
    if "download=1" not in url:
        url = url + ("&" if "?" in url else "?") + "download=1"
    cj = http.cookiejar.CookieJar()
    op = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    op.addheaders = [("User-Agent", UA),
                     ("Accept", "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"),
                     ("Accept-Language", "es-ES,es;q=0.9")]
    with op.open(url, timeout=90) as r:
        data = r.read()
    if data[:2] != b"PK":  # un .xlsx es un zip → empieza con 'PK'
        raise RuntimeError("La descarga no es un Excel (link bloqueado o caducado). Pásame un link nuevo de compartir.")
    d = os.path.dirname(CACHE)
    os.makedirs(d, exist_ok=True)
    # Escribe a temporal y reemplaza atómicamente: siempre queda SOLO el último archivo
    # (el anterior se borra), sin acumular descargas viejas.
    tmp = CACHE + ".tmp"
    with open(tmp, "wb") as f:
        f.write(data)
    os.replace(tmp, CACHE)
    for nombre in os.listdir(d):           # limpieza: deja solo la caché vigente
        p = os.path.join(d, nombre)
        if p != CACHE:
            try:
                os.remove(p)
            except Exception:
                pass
    return CACHE


def wb(forzar=False):
    import openpyxl
    return openpyxl.load_workbook(descargar(forzar), read_only=True, data_only=True)


def cel(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return v


def norm_pat(p):
    return re.sub(r"[^A-Z0-9]", "", str(p or "").upper())


def num(v):
    if isinstance(v, (int, float)):
        return v
    if v is None:
        return 0
    s = re.sub(r"[^\d.\-]", "", str(v))
    try:
        return float(s) if s not in ("", "-", ".") else 0
    except Exception:
        return 0


def fila_header(rows):
    """Detecta la fila de encabezado: la primera (de las 10 primeras) con >=3 textos."""
    for i, r in enumerate(rows[:10]):
        strs = [c for c in r if isinstance(c, str) and c.strip()]
        if len(strs) >= 3:
            return i
    return 0


def leer_hoja(ws, limite=40, buscar=None):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    hi = fila_header(rows)
    header = [(str(c).strip() if c is not None else f"col{j}") for j, c in enumerate(rows[hi])]
    out = []
    b = (buscar or "").lower()
    for r in rows[hi + 1:]:
        if all(c is None or c == "" for c in r):
            continue
        d = {}
        for j, c in enumerate(r):
            if j < len(header) and header[j] and c is not None and c != "":
                d[header[j]] = cel(c)
        if not d:
            continue
        if b and b not in " ".join(str(v) for v in d.values()).lower():
            continue
        out.append(d)
        if len(out) >= limite:
            break
    return header, out


def col(header, *nombres):
    """Índice de la primera columna cuyo nombre contiene alguno de `nombres`."""
    up = [str(h).upper() for h in header]
    for n in nombres:
        for i, h in enumerate(up):
            if n in h:
                return i
    return -1


def cmd_hojas(args):
    libro = wb(args.refrescar)
    hojas = []
    for ws in libro.worksheets:
        rows = list(ws.iter_rows(min_row=1, max_row=12, values_only=True))
        hi = fila_header(rows) if rows else 0
        header = [str(c).strip() for c in (rows[hi] if rows else []) if c is not None and str(c).strip()]
        hojas.append({"hoja": ws.title, "hoja_limpia": ws.title.strip(),
                      "filas": ws.max_row, "columnas": header[:20]})
    salir({"fuente": "Excel global Mallorca", "hojas": hojas})


def _resolver_hoja(libro, nombre):
    """Encuentra la hoja por nombre normalizado (strip + case-insensitive), tolerando
    espacios al inicio/final. Devuelve el nombre REAL (con espacios) o None."""
    objetivo = (nombre or "").strip().casefold()
    for real in libro.sheetnames:      # nombres reales, pueden traer espacios
        if real.strip().casefold() == objetivo:
            return real
    return None


def cmd_hoja(args):
    libro = wb(args.refrescar)
    real = _resolver_hoja(libro, args.nombre)
    if real is None:
        salir({"error": f"Hoja no encontrada: {args.nombre}",
               "hojas_validas": [s.strip() for s in libro.sheetnames]})
    header, filas = leer_hoja(libro[real], limite=min(args.limite, 200), buscar=args.buscar)
    salir({"fuente": "Excel global Mallorca", "hoja": real.strip(), "columnas": header,
           "devueltas": len(filas), "filas": filas})


def cmd_auto(args):
    libro = wb(args.refrescar)
    objetivo = norm_pat(args.patente)
    if not objetivo:
        salir({"error": "Falta --patente"})
    res = {"fuente": "Excel global Mallorca", "patente": args.patente.upper()}
    for clave, hoja in (("stock_valorizado", HOJA_STOCK), ("ventas", HOJA_VENTAS), ("compras", HOJA_COMPRAS)):
        if hoja not in libro.sheetnames:
            continue
        header, filas = leer_hoja(libro[hoja], limite=500)
        ipat = col(header, "PATENTE")
        if ipat < 0:
            continue
        clavep = header[ipat]
        match = [f for f in filas if norm_pat(f.get(clavep)) == objetivo]
        if match:
            res[clave] = match
    if len(res) == 2:
        res["aviso"] = "Sin datos en el Excel para esa patente."
    salir(res)


def cmd_stock(args):
    libro = wb(args.refrescar)
    if HOJA_STOCK not in libro.sheetnames:
        salir({"error": f"No existe la hoja {HOJA_STOCK}"})
    header, filas = leer_hoja(libro[HOJA_STOCK], limite=500)
    ipat, imar, imod = col(header, "PATENTE"), col(header, "MARCA"), col(header, "MODELO")
    icosto, itotal, ipv = col(header, "COSTO"), col(header, "TOTAL"), col(header, "PV")
    hp = header[ipat] if ipat >= 0 else None
    autos, suma_costo, suma_total = [], 0, 0
    for f in filas:
        pat = f.get(hp) if hp else None
        if not norm_pat(pat):
            continue
        costo = num(f.get(header[icosto])) if icosto >= 0 else 0
        total = num(f.get(header[itotal])) if itotal >= 0 else 0
        suma_costo += costo
        suma_total += total
        autos.append({
            "patente": str(pat).upper(),
            "marca": f.get(header[imar]) if imar >= 0 else None,
            "modelo": f.get(header[imod]) if imod >= 0 else None,
            "costo": costo, "total_invertido": total,
            "pv_esperado": num(f.get(header[ipv])) if ipv >= 0 else None,
        })
    salir({"fuente": "Excel global Mallorca", "hoja": HOJA_STOCK, "autos_en_stock": len(autos),
           "costo_total": suma_costo, "total_invertido": suma_total, "autos": autos})


def cmd_ventas(args):
    libro = wb(args.refrescar)
    if HOJA_VENTAS not in libro.sheetnames:
        salir({"error": f"No existe la hoja {HOJA_VENTAS}"})
    header, filas = leer_hoja(libro[HOJA_VENTAS], limite=2000)
    imes, ifec = col(header, "MES"), col(header, "FECHA")
    iventa, imargen = col(header, "VENTA"), col(header, "MARGEN")
    hmes = header[imes] if imes >= 0 else None
    hfec = header[ifec] if ifec >= 0 else None
    filtro = (args.mes or "").strip()
    sel, suma_venta, suma_margen = [], 0, 0
    for f in filas:
        if filtro:
            txt = f"{f.get(hmes) if hmes else ''} {f.get(hfec) if hfec else ''}"
            if filtro.lower() not in str(txt).lower():
                # también probar YYYY-MM contra la fecha
                if not str(f.get(hfec) if hfec else "").startswith(filtro):
                    continue
        venta = num(f.get(header[iventa])) if iventa >= 0 else 0
        margen = num(f.get(header[imargen])) if imargen >= 0 else 0
        suma_venta += venta
        suma_margen += margen
        sel.append(f)
    salir({"fuente": "Excel global Mallorca", "hoja": HOJA_VENTAS,
           "filtro_mes": filtro or None, "ventas_contadas": len(sel),
           "venta_total": suma_venta, "margen_total": suma_margen,
           "detalle": sel[:60]})


def main():
    p = argparse.ArgumentParser()
    sub = p.add_subparsers(dest="cmd")
    for nombre in ("hojas", "auto", "stock", "ventas", "refrescar"):
        sp = sub.add_parser(nombre)
        sp.add_argument("--refrescar", action="store_true")
        if nombre == "auto":
            sp.add_argument("--patente", default="")
        if nombre == "ventas":
            sp.add_argument("--mes", default="")
    sph = sub.add_parser("hoja")
    # No obligatorio: si el nombre llega vacío (hoja con nombre en blanco, que api.py
    # descarta al recortar), el resolver la matchea igual por nombre normalizado.
    sph.add_argument("--nombre", default="")
    sph.add_argument("--limite", type=int, default=40)
    sph.add_argument("--buscar", default="")
    sph.add_argument("--refrescar", action="store_true")

    args = p.parse_args()
    try:
        if args.cmd == "hojas":
            cmd_hojas(args)
        elif args.cmd == "hoja":
            cmd_hoja(args)
        elif args.cmd == "auto":
            cmd_auto(args)
        elif args.cmd == "stock":
            cmd_stock(args)
        elif args.cmd == "ventas":
            cmd_ventas(args)
        elif args.cmd == "refrescar":
            descargar(forzar=True)
            salir({"ok": True, "msg": "Excel re-descargado"})
        else:
            salir({"error": "Comando desconocido",
                   "comandos": ["hojas", "hoja --nombre X", "auto --patente X", "stock", "ventas [--mes YYYY-MM]", "refrescar"]})
    except Exception as e:
        salir({"error": str(e)})


if __name__ == "__main__":
    main()
