#!/usr/bin/env python3
# generar-excel-tag.py — Genera un Excel LISTO "STOCK VALORIZADO + TAG": toma la hoja
# STOCK VALORIZADO del Excel de Mallorca (cache) y agrega una columna TAG (SI/NO) según
# las patentes con tag (historial de correos + leads de Nexus). No toca el OneDrive.
#
#   salida: stock-con-tag.xlsx  (en esta carpeta)

import os, json, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "..", "conector-mallorca", "cache", "mallorca.xlsx")
OUT = os.path.join(BASE, "stock-con-tag.xlsx")

def norm(p): return re.sub(r"[^A-Za-z0-9]", "", str(p or "")).upper()

def patentes_con_tag():
    s = set()
    try:
        d = json.load(open(os.path.join(BASE, "tag-patentes-historicas.json")))
        for p in d.get("patentes", []): s.add(norm(p))
    except Exception: pass
    try:
        d = json.load(open(os.path.join(BASE, "tag-registros.json")))
        for r in d.get("registros", []):
            if r.get("patente") and r.get("estado") != "rechazado": s.add(norm(r["patente"]))
    except Exception: pass
    return s

def main():
    tag = patentes_con_tag()
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["STOCK VALORIZADO"]
    # Columna nueva TAG al final del encabezado (fila 1)
    col = ws.max_column + 1
    hdr = ws.cell(row=1, column=col, value="TAG")
    hdr.font = Font(bold=True, color="FFFFFF")
    hdr.fill = PatternFill("solid", fgColor="2F6DF6")
    hdr.alignment = Alignment(horizontal="center")
    con = 0; total = 0
    verde = PatternFill("solid", fgColor="C9F5D8"); gris = PatternFill("solid", fgColor="F2F2F2")
    for row in range(2, ws.max_row + 1):
        pat = ws.cell(row=row, column=2).value  # col B = PATENTE
        if not pat: continue
        total += 1
        tiene = norm(pat) in tag and len(norm(pat)) >= 5
        c = ws.cell(row=row, column=col, value="SÍ" if tiene else "NO")
        c.alignment = Alignment(horizontal="center")
        c.fill = verde if tiene else gris
        if tiene: con += 1
    wb.save(OUT)
    print(json.dumps({"ok": True, "archivo": OUT, "total": total, "con_tag": con, "sin_tag": total - con}))

if __name__ == "__main__":
    main()
