#!/usr/bin/env python3
# parse-tagtico.py — Lee el Excel mensual de Tag Tico (ANA CLARA - ... .xlsx, hojas por
# año con columnas MONTO, PLACA, ACCION, CLIENTE, FECHA, MARCA, PAGO) y devuelve en JSON
# todas las PLACAS que tuvieron TAG NUEVO o TRASPASO. Es la fuente oficial de "con tag".
#   uso: parse-tagtico.py <ruta.xlsx>

import sys, re, json
import openpyxl

def norm(p): return re.sub(r"[^A-Za-z0-9]", "", str(p or "")).upper()

def main(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    placas = set()
    for sn in wb.sheetnames:
        ws = wb[sn]
        # ubica columnas PLACA y ACCION por encabezado (fila 1)
        hdr = [str(c.value or "").strip().upper() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            ci_placa = hdr.index("PLACA")
        except ValueError:
            ci_placa = 1
        ci_accion = hdr.index("ACCION") if "ACCION" in hdr else 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(ci_placa, ci_accion): continue
            placa = norm(row[ci_placa]); accion = str(row[ci_accion] or "").upper()
            if placa and len(placa) >= 5 and ("TAG" in accion or "TRASPASO" in accion):
                placas.add(placa)
    print(json.dumps({"patentes": sorted(placas)}))

if __name__ == "__main__":
    main(sys.argv[1])
